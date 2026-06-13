"""Spreadsheet round-trip for translations (export / import).

Lets a translator who only has the shipped app — not a dev checkout — edit
ChromIQ's UI language in a spreadsheet.  Every translatable phrase becomes one
row; the translator fills the *translation* column in Excel / LibreOffice /
Sheets and imports the file back.  Imported catalogs are written into the
writable :func:`core.i18n.user_i18n_dir` so the read-only frozen ``.app`` is
never touched, and the i18n loaders prefer that directory.

Two catalogs are covered in one sheet:

* the UI string catalog  ``data/i18n/<code>.json``  (``section = "ui"``); the
  English source string *is* the catalog key.
* the parameter-tooltip overlay  ``data/i18n/parameters.<code>.yaml``
  (``section = "param"``); ``id`` encodes ``tool/flag/field`` (or
  ``tool/flag/labels/<index>`` for the per-choice label lists).

A few ``section = "meta"`` rows carry the target language code, the native
language name (editable) and the ChromIQ version, so the importer can identify
the file and warn on version drift.

This module is pure logic (no Qt) so it is unit-testable headless.  CSV and
XLSX share the single flat-:class:`Row` model.
"""
from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml

from core.i18n import (
    SOURCE_LANGUAGE,
    check_placeholders,
    placeholders,
    user_i18n_dir,
)
from core.logger import get_logger
from core.resource_path import resource_path

try:
    from core.version import APP_VERSION
except Exception:  # pragma: no cover — version module optional in tests
    APP_VERSION = ""

log = get_logger("i18n_roundtrip")

# Column order written to / read from every sheet (first physical row).
COLUMNS = ("section", "id", "english", "translation", "notes")

# Overlay text fields that live directly under tool/flag (mirrors
# core.i18n._PARAM_TEXT_FIELDS); label lists are handled separately.
_PARAM_FIELDS = ("name", "tooltip_title", "tooltip_body")

_META_PREFIX = "@"


@dataclass
class Row:
    section: str          # "meta" | "ui" | "param"
    id: str               # stable locator; for "ui" rows it equals `english`
    english: str          # English source text (the catalog key for UI rows)
    translation: str      # target-language text, blank when untranslated
    notes: str = ""

    def as_tuple(self) -> tuple[str, ...]:
        return (self.section, self.id, self.english, self.translation, self.notes)


@dataclass
class Report:
    """Outcome of validating imported rows before they are written."""
    translated: int = 0
    missing: int = 0
    placeholder_errors: list[str] = field(default_factory=list)
    label_errors: list[str] = field(default_factory=list)
    code_mismatch: str | None = None  # sheet's @target_code, if it differs

    @property
    def has_errors(self) -> bool:
        return bool(self.placeholder_errors or self.label_errors)


# ----------------------------------------------------------------------
# Reading the English source + current translation
# ----------------------------------------------------------------------

def _load_json(path: Path) -> dict[str, Any]:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _ui_source_keys() -> list[str]:
    """All UI catalog keys (English source strings), sorted.

    Taken from the union of the bundled ``data/i18n/*.json`` catalogs, which the
    i18n hygiene tests keep complete and in sync — so any one of them already
    holds the full key set, and the union is robust to a stale file.
    """
    keys: set[str] = set()
    for p in sorted(resource_path("data/i18n").glob("*.json")):
        if p.stem == SOURCE_LANGUAGE:
            continue
        try:
            for k in _load_json(p):
                if not k.startswith(_META_PREFIX):
                    keys.add(k)
        except Exception:
            log.warning("Skipping unreadable catalog %s", p, exc_info=True)
    return sorted(keys)


def _current_ui_catalog(code: str) -> dict[str, str]:
    """The existing ``<code>.json`` map (override-aware), or {} for a new lang."""
    override = user_i18n_dir() / f"{code}.json"
    path = override if override.exists() else resource_path(f"data/i18n/{code}.json")
    if not path.exists():
        return {}
    try:
        return {k: v for k, v in _load_json(path).items() if isinstance(v, str)}
    except Exception:
        log.warning("Unreadable catalog for %r", code, exc_info=True)
        return {}


def _base_parameters() -> dict[str, list[dict[str, Any]]]:
    """``data/parameters.yaml`` → {tool: [param-def, …]} (English source)."""
    with open(resource_path("data/parameters.yaml"), encoding="utf-8") as f:
        return (yaml.safe_load(f) or {}).get("parameters", {})


def _current_param_overlay(code: str) -> dict[str, Any]:
    """Existing ``parameters.<code>.yaml`` overlay (override-aware), nested
    ``{tool: {flag: {field: text, labels: [...]}}}``; {} when absent."""
    override = user_i18n_dir() / f"parameters.{code}.yaml"
    path = override if override.exists() else \
        resource_path(f"data/i18n/parameters.{code}.yaml")
    if not path.exists():
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return (yaml.safe_load(f) or {}).get("parameters", {})
    except Exception:
        log.warning("Unreadable parameter overlay for %r", code, exc_info=True)
        return {}


# ----------------------------------------------------------------------
# Build rows for export
# ----------------------------------------------------------------------

def build_rows(code: str, language_name: str | None = None) -> list[Row]:
    """Every translatable phrase as a flat list of rows for language ``code``.

    Pre-fills the translation column from the current catalog/overlay (blank for
    a brand-new language).  ``language_name`` overrides the stored native name
    (used when the user is starting a new language).
    """
    rows: list[Row] = []

    cur_ui = _current_ui_catalog(code)
    if language_name is None:
        language_name = cur_ui.get("@language_name", "")

    rows.append(Row("meta", "@target_code", "en", code,
                    "Language code — do not change"))
    rows.append(Row("meta", "@chromiq_version", APP_VERSION, "",
                    "Exported from this ChromIQ version"))
    rows.append(Row("meta", "@language_name", "",
                    language_name, "Native name of the language"))

    for key in _ui_source_keys():
        rows.append(Row("ui", key, key, cur_ui.get(key, "")))

    overlay = _current_param_overlay(code)
    for tool, defs in _base_parameters().items():
        tool_tr = overlay.get(tool, {}) if isinstance(overlay, dict) else {}
        for param in defs:
            flag = str(param.get("flag", ""))
            entry_tr = tool_tr.get(flag, {}) if isinstance(tool_tr, dict) else {}
            for fld in _PARAM_FIELDS:
                if fld not in param:
                    continue
                rows.append(Row(
                    "param", f"{tool}/{flag}/{fld}",
                    str(param[fld]),
                    str(entry_tr.get(fld, "")) if isinstance(entry_tr, dict) else "",
                ))
            labels = param.get("labels")
            tr_labels = entry_tr.get("labels") if isinstance(entry_tr, dict) else None
            if isinstance(labels, list):
                for i, lbl in enumerate(labels):
                    cur = ""
                    if isinstance(tr_labels, list) and len(tr_labels) == len(labels):
                        cur = str(tr_labels[i])
                    rows.append(Row(
                        "param", f"{tool}/{flag}/labels/{i}", str(lbl), cur))
    return rows


# ----------------------------------------------------------------------
# Write
# ----------------------------------------------------------------------

def write_csv(rows: list[Row], path: str | Path) -> None:
    # utf-8-sig so Excel opens the file as UTF-8; QUOTE_ALL so the
    # HTML/comma/newline-heavy strings survive a spreadsheet round-trip.
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(COLUMNS)
        for r in rows:
            w.writerow(r.as_tuple())


def write_xlsx(rows: list[Row], path: str | Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "ChromIQ translation"

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    bold = Font(bold=True)
    ws.append(list(COLUMNS))
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill

    wrap = Alignment(wrap_text=True, vertical="top")
    for r in rows:
        ws.append(list(r.as_tuple()))
        # English source is reference-only — grey it so translators edit only
        # the translation column.
        last = ws.max_row
        ws.cell(last, 3).alignment = wrap
        ws.cell(last, 3).font = Font(color="808080")
        ws.cell(last, 4).alignment = wrap

    widths = {"A": 9, "B": 34, "C": 60, "D": 60, "E": 28}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "D2"  # keep section/id/english visible while scrolling
    wb.save(str(path))


# ----------------------------------------------------------------------
# Read
# ----------------------------------------------------------------------

def read_rows(path: str | Path) -> list[Row]:
    """Read a sheet back into rows, auto-detecting CSV vs XLSX by extension."""
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx(path)
    return _read_csv(path)


def _row_from_values(values: list[Any]) -> Row | None:
    cells = [("" if v is None else str(v)) for v in values[:5]]
    cells += [""] * (5 - len(cells))
    section, rid, english, translation, notes = cells
    if not section and not rid:
        return None
    return Row(section.strip(), rid, english, translation, notes)


def _read_csv(path: Path) -> list[Row]:
    rows: list[Row] = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for i, raw in enumerate(reader):
            if i == 0 and raw[:1] == [COLUMNS[0]]:
                continue  # header
            r = _row_from_values(raw)
            if r is not None:
                rows.append(r)
    return rows


def _read_xlsx(path: Path) -> list[Row]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows: list[Row] = []
    for i, raw in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 and (raw[:1] == (COLUMNS[0],) or raw[:1] == [COLUMNS[0]]):
            continue
        r = _row_from_values(list(raw))
        if r is not None:
            rows.append(r)
    wb.close()
    return rows


# ----------------------------------------------------------------------
# Apply (validate + reconstruct catalogs)
# ----------------------------------------------------------------------

def apply_rows(
    code: str, rows: list[Row]
) -> tuple[dict[str, str], dict[str, Any], Report]:
    """Reconstruct the ``<code>.json`` map and ``parameters.<code>.yaml`` overlay
    from imported ``rows`` and validate them.

    Returns ``(json_dict, yaml_dict, report)``.  Entries that fail validation are
    excluded from the returned dicts and listed in ``report``; callers must
    refuse to save when ``report.has_errors``.
    """
    report = Report()
    json_dict: dict[str, str] = {}
    language_name = ""

    # base param defs for label-length validation
    base_params = _base_parameters()
    base_index: dict[tuple[str, str], dict[str, Any]] = {}
    for tool, defs in base_params.items():
        for param in defs:
            base_index[(tool, str(param.get("flag", "")))] = param

    # collect param scalars + label rows
    param_scalars: dict[tuple[str, str], dict[str, str]] = {}
    param_labels: dict[tuple[str, str], dict[int, str]] = {}

    for r in rows:
        if r.section == "meta":
            if r.id == "@language_name" and r.translation.strip():
                language_name = r.translation.strip()
            elif r.id == "@target_code":
                sheet_code = r.translation.strip()
                if sheet_code and sheet_code != code:
                    report.code_mismatch = sheet_code
            continue

        if r.section == "ui":
            key = r.english if r.english else r.id
            tr = r.translation.strip()
            if not tr:
                report.missing += 1
                continue
            if placeholders(key) != placeholders(tr):
                report.placeholder_errors.append(key)
                continue
            json_dict[key] = r.translation
            report.translated += 1
            continue

        if r.section == "param":
            parts = r.id.split("/")
            if len(parts) == 3 and parts[2] in _PARAM_FIELDS:
                tool, flag, fld = parts
                tr = r.translation.strip()
                if not tr:
                    report.missing += 1
                    continue
                if placeholders(r.english) != placeholders(tr):
                    report.placeholder_errors.append(r.id)
                    continue
                param_scalars.setdefault((tool, flag), {})[fld] = r.translation
                report.translated += 1
            elif len(parts) == 4 and parts[2] == "labels":
                tool, flag, _, idx = parts
                try:
                    i = int(idx)
                except ValueError:
                    continue
                param_labels.setdefault((tool, flag), {})[i] = r.translation.strip()

    # belt-and-braces placeholder sweep over the UI dict
    for bad in check_placeholders(json_dict):
        if bad not in report.placeholder_errors:
            report.placeholder_errors.append(bad)
            json_dict.pop(bad, None)

    yaml_dict = _assemble_overlay(
        param_scalars, param_labels, base_index, report)

    json_dict = {"@language_name": language_name, **json_dict} if language_name \
        else json_dict
    return json_dict, yaml_dict, report


def _assemble_overlay(
    scalars: dict[tuple[str, str], dict[str, str]],
    labels: dict[tuple[str, str], dict[int, str]],
    base_index: dict[tuple[str, str], dict[str, Any]],
    report: Report,
) -> dict[str, Any]:
    """Build the nested ``{tool: {flag: {...}}}`` overlay, validating that any
    translated label list is complete and matches the English length (a partial
    or wrong-length list would desync from its `choices`, so it is dropped and
    reported)."""
    overlay: dict[str, dict[str, dict[str, Any]]] = {}

    for (tool, flag), fields in scalars.items():
        overlay.setdefault(tool, {})[flag] = dict(fields)

    for (tool, flag), idx_map in labels.items():
        base = base_index.get((tool, flag), {})
        base_labels = base.get("labels")
        if not isinstance(base_labels, list):
            continue
        translated = {i: v for i, v in idx_map.items() if v}
        if not translated:
            continue  # no label translated — leave English labels in place
        if len(translated) != len(base_labels):
            report.label_errors.append(f"{tool}/{flag}/labels")
            continue
        ordered = [translated[i] for i in range(len(base_labels))]
        overlay.setdefault(tool, {}).setdefault(flag, {})["labels"] = ordered
        report.translated += len(ordered)

    return {"parameters": overlay} if overlay else {}


# ----------------------------------------------------------------------
# Save into the writable user override dir
# ----------------------------------------------------------------------

def save_translation(
    code: str, json_dict: dict[str, str], yaml_dict: dict[str, Any]
) -> Path:
    """Write ``<code>.json`` (and ``parameters.<code>.yaml`` if non-empty) into
    :func:`core.i18n.user_i18n_dir`.  Returns the directory written to."""
    out = user_i18n_dir()
    out.mkdir(parents=True, exist_ok=True)

    with open(out / f"{code}.json", "w", encoding="utf-8") as f:
        json.dump(json_dict, f, ensure_ascii=False, indent=1, sort_keys=True)
        f.write("\n")

    yaml_path = out / f"parameters.{code}.yaml"
    if yaml_dict.get("parameters"):
        with open(yaml_path, "w", encoding="utf-8") as f:
            yaml.safe_dump(yaml_dict, f, allow_unicode=True, sort_keys=True,
                           default_flow_style=False, width=10_000)
    elif yaml_path.exists():
        yaml_path.unlink()  # a re-import that cleared all params removes the stale overlay

    log.info("Saved translation %r → %s", code, out)
    return out
